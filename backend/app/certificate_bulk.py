"""
Bulk internship certificate upload (admin).
Template → validate → create records → background PDF generation + emails.
"""
from __future__ import annotations

import csv
import io
import logging
import re
import threading
from datetime import datetime, date
from typing import Any

from bson import Binary, ObjectId
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.certificate_storage import save_certificate_pdf
from app.certificate_verification import (
    allocate_certificate_number,
    certificate_pdf_bytes,
    find_certificate_by_no,
    normalize_cert_no,
    parse_certificate_admin_fields,
    verify_url_for_cert,
)
from app.db import (
    get_bulk_certificate_jobs_collection,
    get_certificates_collection,
    get_courses_collection,
)

logger = logging.getLogger(__name__)

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 1000
BATCH_TICK = 8

MODE_ALLOWED = {"online", "offline", "hybrid"}
RATING_ALLOWED = {"excellent", "very good", "good", "average"}

# Exact headers from product spec (order matters for template).
TEMPLATE_HEADERS = [
    "Certificate Number",
    "Student Name",
    "College Name",
    "Course",
    "Branch",
    "Semester",
    "Registration No",
    "Session",
    "Domain",
    "Duration",
    "Mode",
    "Start Date",
    "End Date",
    "Marks",
    "Attendance",
    "Performance Rating",
]

# Optional extra column for student email delivery (Section 5).
OPTIONAL_HEADERS = ["Student Email"]

HEADER_ALIASES = {
    "certificate number": "Certificate Number",
    "cert no": "Certificate Number",
    "certno": "Certificate Number",
    "student name": "Student Name",
    "name": "Student Name",
    "college name": "College Name",
    "college": "College Name",
    "course": "Course",
    "branch": "Branch",
    "semester": "Semester",
    "registration no": "Registration No",
    "registration number": "Registration No",
    "session": "Session",
    "domain": "Domain",
    "duration": "Duration",
    "mode": "Mode",
    "start date": "Start Date",
    "end date": "End Date",
    "marks": "Marks",
    "attendance": "Attendance",
    "performance rating": "Performance Rating",
    "student email": "Student Email",
    "email": "Student Email",
}

INSTRUCTIONS = [
    "Do not delete or rename any column header.",
    "Leave Certificate Number empty if you want the system to generate it automatically.",
    "Date format must be DD-MM-YYYY (e.g. 15-08-2026).",
    "Mode allowed values: Online, Offline, Hybrid.",
    "Performance Rating allowed values: Excellent, Very Good, Good, Average.",
    "Do not upload more than 1000 rows in one file.",
    "Domain must match an existing training title (case-insensitive).",
    "Student Email (optional column) — if provided, the student receives the certificate PDF by email.",
]


def _norm_header(h: str) -> str:
    key = re.sub(r"\s+", " ", (h or "").strip().lower())
    return HEADER_ALIASES.get(key, (h or "").strip())


def training_titles() -> list[str]:
    titles: list[str] = []
    try:
        for c in get_courses_collection().find({"active": True}, {"title": 1}):
            t = (c.get("title") or "").strip()
            if t:
                titles.append(t)
    except Exception as exc:
        logger.warning("training_titles failed: %s", exc)
    return titles


def _match_domain(domain: str, titles: list[str]) -> str | None:
    d = (domain or "").strip().lower()
    if not d:
        return None
    for t in titles:
        if t.strip().lower() == d:
            return t
    return None


def _parse_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    # Excel serial sometimes arrives as int/float via openpyxl as datetime already
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s[:10] if fmt.startswith("%Y") else s, fmt).date()
        except ValueError:
            continue
    # try longer strings
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _date_to_store(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def build_xlsx_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificates"
    for col, h in enumerate(TEMPLATE_HEADERS + OPTIONAL_HEADERS, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
    ws2 = wb.create_sheet("Instructions")
    ws2.cell(1, 1, "Bulk Certificate Upload — Instructions").font = Font(bold=True)
    for i, line in enumerate(INSTRUCTIONS, start=3):
        ws2.cell(i, 1, f"• {line}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_template() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS + OPTIONAL_HEADERS)
    w.writerow([])  # blank sample
    return ("\ufeff" + buf.getvalue()).encode("utf-8")  # BOM for Excel


def _rows_from_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_norm_header(str(h) if h is not None else "") for h in header_row]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            if isinstance(val, datetime):
                d[h] = val.strftime("%d-%m-%Y")
            elif isinstance(val, date):
                d[h] = val.strftime("%d-%m-%Y")
            else:
                d[h] = "" if val is None else str(val).strip()
        out.append(d)
    return headers, out


def _rows_from_csv(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return [], []
    headers = [_norm_header(h) for h in header_row]
    out: list[dict[str, str]] = []
    for row in reader:
        if not row or all(not str(c).strip() for c in row):
            continue
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = row[i].strip() if i < len(row) else ""
        out.append(d)
    return headers, out


def parse_upload_file(raw: bytes, filename: str) -> tuple[list[dict[str, str]] | None, str | None]:
    """Returns (rows, error_message)."""
    name = (filename or "").lower()
    if len(raw) > MAX_FILE_BYTES:
        return None, "File is larger than 10 MB."
    if not raw:
        return None, "The file is empty."
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        return None, "Legacy .xls is not supported. Please save as .xlsx or .csv and try again."
    try:
        if name.endswith(".csv"):
            headers, rows = _rows_from_csv(raw)
        elif name.endswith(".xlsx"):
            headers, rows = _rows_from_xlsx(raw)
        else:
            return None, "Accepted formats: .xlsx or .csv."
    except Exception as exc:
        logger.exception("parse upload failed")
        return None, f"Could not read file: {exc}"

    if not rows:
        return None, "The file has no data rows."
    if len(rows) > MAX_ROWS:
        return None, f"File has {len(rows)} rows; maximum is {MAX_ROWS}."

    missing = [h for h in TEMPLATE_HEADERS if h not in headers]
    if missing:
        return None, f"Missing required column headers: {', '.join(missing)}"
    return rows, None


def validate_bulk_rows(rows: list[dict[str, str]]) -> dict[str, Any]:
    titles = training_titles()
    title_set_empty = len(titles) == 0
    seen_in_file: set[str] = set()
    preview: list[dict[str, Any]] = []
    valid_count = 0
    error_count = 0

    for idx, row in enumerate(rows, start=2):  # Excel row (header=1)
        errors: list[str] = []
        student = (row.get("Student Name") or "").strip()
        if not student:
            errors.append("Student name missing")

        domain_raw = (row.get("Domain") or "").strip()
        domain_matched = _match_domain(domain_raw, titles) if domain_raw else None
        if not domain_raw:
            errors.append("Domain missing")
        elif title_set_empty:
            errors.append("No active trainings found to match Domain")
        elif not domain_matched:
            errors.append("Domain does not match an existing training title")

        duration = (row.get("Duration") or "").strip()
        if not duration:
            errors.append("Duration missing")

        mode = (row.get("Mode") or "").strip()
        if not mode:
            errors.append("Mode missing")
        elif mode.lower() not in MODE_ALLOWED:
            errors.append("Mode must be Online, Offline, or Hybrid")

        start = _parse_date(row.get("Start Date"))
        end = _parse_date(row.get("End Date"))
        if not start:
            errors.append("Start Date invalid (use DD-MM-YYYY)")
        if not end:
            errors.append("End Date invalid (use DD-MM-YYYY)")
        if start and end and end < start:
            errors.append("End Date must be on or after Start Date")

        marks_raw = (row.get("Marks") or "").strip()
        if marks_raw:
            try:
                m = float(marks_raw)
                if m < 0 or m > 100:
                    errors.append("Marks must be between 0 and 100")
            except ValueError:
                errors.append("Marks must be a number")

        sem_raw = (row.get("Semester") or "").strip()
        if sem_raw:
            try:
                sem = int(float(sem_raw))
                if sem < 1 or sem > 8:
                    errors.append("Semester must be between 1 and 8")
            except ValueError:
                errors.append("Semester must be a number")

        rating = (row.get("Performance Rating") or "").strip()
        if rating and rating.lower() not in RATING_ALLOWED:
            errors.append("Performance Rating must be Excellent, Very Good, Good, or Average")

        cert_no = normalize_cert_no(row.get("Certificate Number") or "")
        if cert_no:
            if cert_no in seen_in_file:
                errors.append("Duplicate certificate number in this file")
            else:
                seen_in_file.add(cert_no)
            if find_certificate_by_no(cert_no):
                errors.append("Certificate number already exists")

        ok = len(errors) == 0
        if ok:
            valid_count += 1
        else:
            error_count += 1

        # Normalized payload for generate step
        payload = {
            "certNo": cert_no,
            "autoGenerateCertNo": not bool(cert_no),
            "studentName": student,
            "studentEmail": (row.get("Student Email") or "").strip(),
            "collegeName": (row.get("College Name") or "").strip(),
            "course": (row.get("Course") or "").strip(),
            "branch": (row.get("Branch") or "").strip(),
            "semester": sem_raw,
            "registrationNo": (row.get("Registration No") or "").strip(),
            "session": (row.get("Session") or "").strip(),
            "domain": domain_matched or domain_raw,
            "duration": duration,
            "mode": mode.title() if mode.lower() in MODE_ALLOWED else mode,
            "internshipStartDate": _date_to_store(start) if start else "",
            "internshipEndDate": _date_to_store(end) if end else "",
            "marks": marks_raw,
            "attendance": (row.get("Attendance") or "").strip(),
            "performanceRating": rating or "Good",
        }

        preview.append({
            "row": idx,
            "status": "valid" if ok else "error",
            "errors": errors,
            "errorReason": "; ".join(errors) if errors else "",
            "studentName": student,
            "domain": domain_matched or domain_raw,
            "certNo": cert_no or "(auto)",
            "payload": payload if ok else None,
            "raw": {h: row.get(h, "") for h in TEMPLATE_HEADERS + OPTIONAL_HEADERS},
        })

    return {
        "total": len(preview),
        "validCount": valid_count,
        "errorCount": error_count,
        "rows": preview,
        "trainingTitles": titles,
    }


def build_errors_xlsx(preview_rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    headers = TEMPLATE_HEADERS + OPTIONAL_HEADERS + ["Error"]
    for col, h in enumerate(headers, start=1):
        ws.cell(1, col, h).font = Font(bold=True)
    r = 2
    for item in preview_rows:
        if item.get("status") != "error":
            continue
        raw = item.get("raw") or {}
        for col, h in enumerate(TEMPLATE_HEADERS + OPTIONAL_HEADERS, start=1):
            ws.cell(r, col, raw.get(h, ""))
        ws.cell(r, len(TEMPLATE_HEADERS) + len(OPTIONAL_HEADERS) + 1, item.get("errorReason") or "")
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _serialize_job(job: dict) -> dict:
    created = job.get("createdAt")
    updated = job.get("updatedAt")
    completed = job.get("completedAt")
    return {
        "id": str(job.get("_id", "")),
        "adminName": job.get("adminName") or "",
        "adminEmail": job.get("adminEmail") or "",
        "fileName": job.get("fileName") or "",
        "totalRows": int(job.get("totalRows") or 0),
        "validRows": int(job.get("validRows") or 0),
        "errorRows": int(job.get("errorRows") or 0),
        "createdCount": int(job.get("createdCount") or 0),
        "pdfDone": int(job.get("pdfDone") or 0),
        "pdfFailed": int(job.get("pdfFailed") or 0),
        "status": job.get("status") or "",
        "createdAt": created.strftime("%Y-%m-%d %H:%M UTC") if hasattr(created, "strftime") else str(created or ""),
        "updatedAt": updated.strftime("%Y-%m-%d %H:%M UTC") if hasattr(updated, "strftime") else str(updated or ""),
        "completedAt": completed.strftime("%Y-%m-%d %H:%M UTC") if hasattr(completed, "strftime") else str(completed or ""),
        "certificateIds": [str(x) for x in (job.get("certificateIds") or [])],
        "message": job.get("message") or "",
    }


def create_bulk_certificate_job(
    *,
    app,
    admin_id: str,
    admin_email: str,
    admin_name: str,
    file_name: str,
    file_bytes: bytes | None,
    valid_payloads: list[dict],
    total_rows: int,
    error_rows: int,
) -> dict:
    coll = get_certificates_collection()
    job_coll = get_bulk_certificate_jobs_collection()
    now = datetime.utcnow()
    cert_ids: list[str] = []
    created_payloads: list[dict] = []

    for p in valid_payloads:
        fields = parse_certificate_admin_fields(p)
        cert_no = fields.get("certNo") or ""
        if not cert_no or p.get("autoGenerateCertNo"):
            cert_no = allocate_certificate_number(fields.get("domain") or "INT")
            fields["certNo"] = cert_no
        if find_certificate_by_no(cert_no):
            # race / collision — skip with new number once
            cert_no = allocate_certificate_number(fields.get("domain") or "INT")
            fields["certNo"] = cert_no

        doc = {
            **{k: v for k, v in fields.items() if v or k == "certNo"},
            "certNo": cert_no,
            "status": "valid",
            "pdfStatus": "pending",
            "source": "admin-bulk",
            "bulkUploadedAt": now,
            "issueDate": now,
            "createdAt": now,
            "updatedAt": now,
        }
        if not doc.get("completionDate") and doc.get("internshipEndDate"):
            doc["completionDate"] = doc["internshipEndDate"]
        res = coll.insert_one(doc)
        cid = str(res.inserted_id)
        cert_ids.append(cid)
        created_payloads.append({"id": cid, "certNo": cert_no, "email": fields.get("studentEmail") or ""})

    job_doc: dict[str, Any] = {
        "adminId": admin_id,
        "adminEmail": (admin_email or "").strip().lower(),
        "adminName": admin_name or "",
        "fileName": file_name or "",
        "totalRows": total_rows,
        "validRows": len(valid_payloads),
        "errorRows": error_rows,
        "createdCount": len(cert_ids),
        "certificateIds": cert_ids,
        "pdfDone": 0,
        "pdfFailed": 0,
        "status": "in_progress",
        "createdAt": now,
        "updatedAt": now,
        "message": f"{len(cert_ids)} certificates queued for PDF generation.",
    }
    if file_bytes:
        job_doc["originalFile"] = Binary(file_bytes)
        job_doc["originalFileSize"] = len(file_bytes)

    res = job_coll.insert_one(job_doc)
    job_doc["_id"] = res.inserted_id

    # Link certs to job
    if cert_ids:
        coll.update_many(
            {"_id": {"$in": [ObjectId(i) for i in cert_ids if ObjectId.is_valid(i)]}},
            {"$set": {"bulkJobId": str(res.inserted_id)}},
        )

    _spawn_bulk_cert_job(app, str(res.inserted_id))
    return job_doc


def _spawn_bulk_cert_job(app, job_id: str) -> None:
    def run():
        with app.app_context():
            try:
                process_bulk_cert_job(job_id, max_items=10_000)
            except Exception:
                logger.exception("bulk cert job failed job=%s", job_id)

    t = threading.Thread(target=run, daemon=True, name=f"bulk-cert-{job_id[:8]}")
    t.start()


def process_bulk_cert_job(job_id: str, *, max_items: int = BATCH_TICK) -> dict | None:
    """Generate PDFs for pending certs in a job. Safe to call repeatedly (poll/tick)."""
    from flask import current_app
    from app.email_smtp import send_email
    from app.notifications import schedule_certificate_email

    if not ObjectId.is_valid(job_id):
        return None
    job_coll = get_bulk_certificate_jobs_collection()
    cert_coll = get_certificates_collection()
    job = job_coll.find_one({"_id": ObjectId(job_id)})
    if not job:
        return None

    ids = [ObjectId(x) for x in (job.get("certificateIds") or []) if ObjectId.is_valid(str(x))]
    pending = list(
        cert_coll.find({"_id": {"$in": ids}, "pdfStatus": "pending"}).limit(max_items)
    )
    app = current_app._get_current_object()  # type: ignore[attr-defined]

    for c in pending:
        try:
            pdf = certificate_pdf_bytes(c)
            key = save_certificate_pdf(pdf, cert_no=str(c.get("certNo") or ""))
            cert_coll.update_one(
                {"_id": c["_id"]},
                {
                    "$set": {
                        "certificatePdfKey": key,
                        "pdfStatus": "generated",
                        "pdfGeneratedAt": datetime.utcnow(),
                        "updatedAt": datetime.utcnow(),
                    },
                    "$unset": {"pdfError": ""},
                },
            )
            job_coll.update_one({"_id": job["_id"]}, {"$inc": {"pdfDone": 1}, "$set": {"updatedAt": datetime.utcnow()}})
            email = (c.get("studentEmail") or "").strip()
            if email:
                schedule_certificate_email(
                    app,
                    c.get("studentName") or "Student",
                    email,
                    c.get("domain") or c.get("programName") or "Internship",
                    str(c.get("certNo") or ""),
                    pdf,
                )
        except Exception as exc:
            logger.exception("PDF gen failed cert=%s", c.get("certNo"))
            cert_coll.update_one(
                {"_id": c["_id"]},
                {
                    "$set": {
                        "pdfStatus": "failed",
                        "pdfError": str(exc)[:500],
                        "updatedAt": datetime.utcnow(),
                    }
                },
            )
            job_coll.update_one({"_id": job["_id"]}, {"$inc": {"pdfFailed": 1}, "$set": {"updatedAt": datetime.utcnow()}})

    job = job_coll.find_one({"_id": ObjectId(job_id)})
    if not job:
        return None

    remaining = cert_coll.count_documents({"_id": {"$in": ids}, "pdfStatus": "pending"})
    if remaining == 0 and job.get("status") == "in_progress":
        job_coll.update_one(
            {"_id": job["_id"]},
            {
                "$set": {
                    "status": "completed",
                    "completedAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow(),
                    "message": (
                        f"Bulk certificate generation complete. "
                        f"{int(job.get('pdfDone') or 0)} ready, {int(job.get('pdfFailed') or 0)} failed."
                    ),
                }
            },
        )
        admin_email = (job.get("adminEmail") or "").strip()
        if admin_email:
            try:
                send_email(
                    current_app.config,
                    admin_email,
                    "Bulk certificate generation complete",
                    (
                        f"<p>Bulk certificate generation complete.</p>"
                        f"<p><strong>{int(job.get('pdfDone') or 0)}</strong> certificates ready "
                        f"({int(job.get('pdfFailed') or 0)} failed).</p>"
                        f"<p>File: {job.get('fileName') or ''}</p>"
                    ),
                    text_body=(
                        f"Bulk certificate generation complete.\n\n"
                        f"File: {job.get('fileName') or ''}\n"
                        f"Created: {job.get('createdCount') or 0}\n"
                        f"PDFs ready: {job.get('pdfDone') or 0}\n"
                        f"PDF failures: {job.get('pdfFailed') or 0}\n"
                    ),
                )
            except Exception:
                logger.exception("admin completion email failed")
        job = job_coll.find_one({"_id": ObjectId(job_id)})

    return job


def retry_certificate_pdf(cert_id: str, *, app) -> tuple[dict | None, str | None]:
    from app.notifications import schedule_certificate_email

    if not ObjectId.is_valid(cert_id):
        return None, "Invalid certificate id"
    coll = get_certificates_collection()
    c = coll.find_one({"_id": ObjectId(cert_id)})
    if not c:
        return None, "Certificate not found"
    was_failed = (c.get("pdfStatus") or "") == "failed"
    try:
        pdf = certificate_pdf_bytes(c)
        key = save_certificate_pdf(pdf, cert_no=str(c.get("certNo") or ""))
        coll.update_one(
            {"_id": c["_id"]},
            {
                "$set": {
                    "certificatePdfKey": key,
                    "pdfStatus": "generated",
                    "pdfGeneratedAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow(),
                },
                "$unset": {"pdfError": ""},
            },
        )
        email = (c.get("studentEmail") or "").strip()
        if email:
            schedule_certificate_email(
                app,
                c.get("studentName") or "Student",
                email,
                c.get("domain") or c.get("programName") or "Internship",
                str(c.get("certNo") or ""),
                pdf,
            )
        jid = str(c.get("bulkJobId") or "")
        if jid and ObjectId.is_valid(jid) and was_failed:
            get_bulk_certificate_jobs_collection().update_one(
                {"_id": ObjectId(jid), "pdfFailed": {"$gt": 0}},
                {"$inc": {"pdfDone": 1, "pdfFailed": -1}, "$set": {"updatedAt": datetime.utcnow()}},
            )
        updated = coll.find_one({"_id": c["_id"]})
        return updated, None
    except Exception as exc:
        coll.update_one(
            {"_id": c["_id"]},
            {"$set": {"pdfStatus": "failed", "pdfError": str(exc)[:500], "updatedAt": datetime.utcnow()}},
        )
        return None, str(exc)


def list_bulk_jobs(*, limit: int = 50) -> list[dict]:
    coll = get_bulk_certificate_jobs_collection()
    out = []
    for job in coll.find({}, {"originalFile": 0}).sort("createdAt", -1).limit(limit):
        out.append(_serialize_job(job))
    return out


def get_bulk_job(job_id: str, *, include_file: bool = False) -> dict | None:
    if not ObjectId.is_valid(job_id):
        return None
    proj = None if include_file else {"originalFile": 0}
    job = get_bulk_certificate_jobs_collection().find_one({"_id": ObjectId(job_id)}, proj)
    return job


def active_bulk_progress() -> dict | None:
    """Latest in-progress job for banner."""
    job = get_bulk_certificate_jobs_collection().find_one(
        {"status": "in_progress"},
        {"originalFile": 0},
        sort=[("createdAt", -1)],
    )
    if not job:
        return None
    # Tick a few PDFs while polling (helps Lambda)
    try:
        from flask import current_app

        process_bulk_cert_job(str(job["_id"]), max_items=BATCH_TICK)
        job = get_bulk_certificate_jobs_collection().find_one({"_id": job["_id"]}, {"originalFile": 0})
    except Exception:
        logger.exception("active progress tick failed")
    if not job:
        return None
    return _serialize_job(job)
