"""Build and parse enrollment / certificate workflow Excel workbooks (openpyxl)."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

def _assign_col_title(a: dict, index: int) -> str:
    t = (a.get("title") or f"Assignment_{index + 1}").strip()[:120]
    safe = re.sub(r"[^\w\s-]", "_", t, flags=re.UNICODE)
    safe = re.sub(r"[\s_]+", "_", safe).strip("_")[:60] or f"A{index + 1}"
    return f"Assign_{safe}_Submitted"


def merged_student_fields_for_admin(e: dict, u: Optional[dict]) -> Dict[str, str]:
    """Merge user profile with enrollment certificateProfile (checkout) for display and export."""
    prof = e.get("certificateProfile") if isinstance(e.get("certificateProfile"), dict) else {}
    u = u or {}

    def first(*vals) -> str:
        for v in vals:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
        return ""

    return {
        "name": first(u.get("name"), u.get("fullName"), prof.get("fullName")),
        "email": first(u.get("email"), prof.get("email")),
        "mobile": first(u.get("mobile"), prof.get("mobile"), prof.get("phone")),
        "university": first(u.get("university"), prof.get("university")),
        "collegeName": first(u.get("collegeName"), prof.get("collegeName")),
        "course": first(u.get("course"), prof.get("course")),
        "branch": first(u.get("stream"), u.get("branch"), prof.get("branchOrSubject"), prof.get("stream")),
        "semester": first(u.get("semester"), prof.get("semester")),
        "registrationNo": first(u.get("collegeRegNo"), prof.get("registrationNumber"), prof.get("collegeRegNo")),
    }


def assignment_submissions_submitted_count(e: dict) -> int:
    subs = e.get("assignmentSubmissions") if isinstance(e.get("assignmentSubmissions"), list) else []
    return sum(1 for s in subs if isinstance(s, dict) and s.get("submittedAt"))


def enrollment_created_date_str(e: dict) -> str:
    ca = e.get("createdAt")
    if ca is None:
        return ""
    if hasattr(ca, "strftime"):
        return ca.strftime("%Y-%m-%d")
    return str(ca)[:10]


def enrollment_sheet_headers(course: dict) -> List[str]:
    base = [
        "EnrollmentId",
        "UserId",
        "Email",
        "Name",
        "Mobile",
        "University",
        "College",
        "Course",
        "Branch",
        "Semester",
        "RegistrationNo",
        "EnrolledDate",
        "SubmissionsCount",
        "CompletionQuizPassed",
        "CompletionQuizScore",
        "CertificateIssued",
        "CertificateNumber",
        "ApproveCertificate",
    ]
    assigns = course.get("assignments") or []
    if isinstance(assigns, list):
        for i, a in enumerate(assigns):
            if isinstance(a, dict):
                base.append(_assign_col_title(a, i))
    return base


def build_enrollment_workbook_bytes(course: dict, rows: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook

    headers = enrollment_sheet_headers(course)
    wb = Workbook()
    ws = wb.active
    ws.title = "Enrollments"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_workbook_rows(file_stream) -> Tuple[List[str], List[Dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c or "").strip() for c in header_row]
    out: List[Dict[str, Any]] = []
    for r in rows_iter:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    return headers, out


def norm_approve_certificate_cell(val: Any) -> bool:
    s = str(val or "").strip().upper()
    return s in ("Y", "YES", "TRUE", "1")


def row_enrollment_id(row: Dict[str, Any]) -> Optional[str]:
    from bson import ObjectId

    for key in ("EnrollmentId", "enrollmentId", "enrollment_id"):
        if key in row and row[key] is not None:
            s = str(row[key]).strip()
            if s and ObjectId.is_valid(s):
                return s
    return None


def row_email(row: Dict[str, Any]) -> str:
    for key in ("Email", "email"):
        if key in row and row[key] is not None:
            return str(row[key]).strip().lower()
    return ""


def export_row_for_enrollment(course: dict, e: dict, u: Optional[dict]) -> Dict[str, Any]:
    pq = e.get("pythonQuiz") or {}
    cc = e.get("courseCertificate") or {}
    subs = e.get("assignmentSubmissions") if isinstance(e.get("assignmentSubmissions"), list) else []
    m = merged_student_fields_for_admin(e, u)
    by_aid: Dict[str, str] = {}
    for s in subs:
        if isinstance(s, dict):
            aid = str(s.get("assignmentId") or "").strip()
            by_aid[aid] = "Y" if s.get("submittedAt") else "N"
    row: Dict[str, Any] = {
        "EnrollmentId": str(e["_id"]),
        "UserId": str(e.get("userId") or ""),
        "Email": m["email"],
        "Name": m["name"],
        "Mobile": m["mobile"],
        "University": m["university"],
        "College": m["collegeName"],
        "Course": m["course"],
        "Branch": m["branch"],
        "Semester": m["semester"],
        "RegistrationNo": m["registrationNo"],
        "EnrolledDate": enrollment_created_date_str(e),
        "SubmissionsCount": assignment_submissions_submitted_count(e),
        "CompletionQuizPassed": "Y" if pq.get("passedAt") else "N",
        "CompletionQuizScore": pq.get("scorePercent") if pq.get("scorePercent") is not None else "",
        "CertificateIssued": "Y" if (cc.get("certNo") or "").strip() else "N",
        "CertificateNumber": (cc.get("certNo") or "").strip(),
        "ApproveCertificate": "N",
    }
    assigns = course.get("assignments") or []
    if isinstance(assigns, list):
        for i, a in enumerate(assigns):
            if not isinstance(a, dict):
                continue
            col = _assign_col_title(a, i)
            aid = str(a.get("id") or "").strip()
            if not aid:
                aid = f"idx_{i}"
            row[col] = by_aid.get(aid, "N")
    return row
